"""Price History Pivot — Streamlit app.

Upload a flat sales-lines Excel export and get one row per item No., with every
transaction laid out newest -> oldest as repeated [R, Date, Qty, Price, Curr.] blocks.

Access is gated by a password stored in Streamlit secrets (see README).
"""

from __future__ import annotations

import hmac
import io
import re
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BLOCK = ["R", "Date", "Qty", "Price", "Curr."]

# Harb Electric brand palette, sampled from harbelectric.com.
BRAND_BLUE = "#005AA7"
BRAND_BLUE_DARK = "#00447E"
BRAND_INK = "#16171E"
BRAND_GREY = "#626974"
BRAND_LINE = "#E3E7EC"
BRAND_SURFACE = "#FFFFFF"
BRAND_CANVAS = "#F4F6F9"

# Fill for the R cells that actually hold an occurrence number. Change here to
# restyle both the Excel export and the on-screen preview.
R_BLUE = BRAND_BLUE.lstrip("#")

# Header aliases, matched after normalising (lowercase, alphanumerics only).
FIELDS = {
    "no": ["no", "nos", "itemno", "item", "itemnumber", "partno", "partnumber"],
    "desc": ["description", "itemdescription", "desc"],
    "qty": ["quantity", "qty"],
    "price": ["ocnetprice", "netpriceoc", "ocnet", "priceoc"],
    "curr": ["currency", "curr", "currencycode"],
    "date": ["postingdate", "postingdt", "date", "documentdate"],
}
LABELS = {
    "no": "No.",
    "qty": "Quantity",
    "price": "OC Net Price",
    "curr": "Currency",
    "date": "Posting Date",
}

st.set_page_config(
    page_title="Price History Pivot · Harb Electric",
    page_icon="data:image/svg+xml,"
    "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'>"
    "<rect width='32' height='32' rx='7' fill='%23005AA7'/>"
    "<path d='M18.6 4 9 18h5.4L13 28l9.6-14h-5.4z' fill='white'/></svg>",
    layout="wide",
)


# -------------------------------------------------------------------------- chrome

# Inline mark: a bolt in a rounded square. Swap in the official Harb Electric
# logo file here if you'd rather use the real asset.
LOGO_SVG = (
    '<svg width="38" height="38" viewBox="0 0 32 32" role="img" aria-label="Harb Electric">'
    f'<rect width="32" height="32" rx="7" fill="{BRAND_BLUE}"/>'
    '<path d="M18.6 4 9 18h5.4L13 28l9.6-14h-5.4z" fill="#fff"/></svg>'
)

CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&family=Teko:wght@500;600&display=swap');

:root {{
  --brand: {BRAND_BLUE};
  --brand-dark: {BRAND_BLUE_DARK};
  --ink: {BRAND_INK};
  --grey: {BRAND_GREY};
  --line: {BRAND_LINE};
  --surface: {BRAND_SURFACE};
  --canvas: {BRAND_CANVAS};
}}

html, body, [class*="st-"], .stMarkdown, input, button, select, textarea {{
  font-family: 'Barlow', -apple-system, 'Segoe UI', Roboto, sans-serif;
}}
[data-testid="stAppViewContainer"] {{ background: var(--canvas); }}
[data-testid="stHeader"] {{ background: transparent; }}
.block-container {{ padding-top: 1.6rem; max-width: 1280px; }}

/* ---- masthead ---- */
.hb-head {{
  display: flex; align-items: center; gap: .85rem;
  padding: 1rem 1.35rem; margin-bottom: 1.4rem;
  background: var(--surface); border: 1px solid var(--line);
  border-radius: 12px; border-top: 3px solid var(--brand);
  box-shadow: 0 1px 2px rgba(22,23,30,.05);
}}
.hb-head .hb-title {{
  font-family: 'Teko', 'Barlow', sans-serif; font-weight: 600;
  font-size: 1.9rem; line-height: 1; letter-spacing: .02em;
  color: var(--ink); margin: 0;
}}
.hb-head .hb-sub {{
  font-size: .82rem; color: var(--grey); margin: .15rem 0 0;
  letter-spacing: .04em; text-transform: uppercase;
}}
.hb-head .hb-spacer {{ flex: 1 1 auto; }}
.hb-badge {{
  font-size: .72rem; font-weight: 600; letter-spacing: .06em; text-transform: uppercase;
  color: var(--brand); background: rgba(0,90,167,.08);
  border: 1px solid rgba(0,90,167,.18); border-radius: 999px; padding: .3rem .7rem;
}}

/* ---- section labels ---- */
.hb-step {{
  display: flex; align-items: center; gap: .55rem;
  font-size: .78rem; font-weight: 700; letter-spacing: .1em; text-transform: uppercase;
  color: var(--grey); margin: .2rem 0 .6rem;
}}
.hb-step span.n {{
  display: inline-grid; place-items: center; width: 1.35rem; height: 1.35rem;
  border-radius: 50%; background: var(--brand); color: #fff; font-size: .72rem;
}}

/* ---- stat cards ---- */
.hb-stats {{ display: flex; flex-wrap: wrap; gap: .9rem; margin: .2rem 0 1.1rem; }}
.hb-stat {{
  flex: 1 1 170px; background: var(--surface); border: 1px solid var(--line);
  border-radius: 10px; padding: .85rem 1rem; border-left: 3px solid var(--brand);
}}
.hb-stat .v {{
  font-family: 'Teko','Barlow',sans-serif; font-size: 2rem; line-height: 1.05;
  color: var(--ink); font-weight: 600; font-variant-numeric: tabular-nums;
}}
.hb-stat .k {{
  font-size: .74rem; letter-spacing: .08em; text-transform: uppercase; color: var(--grey);
}}

/* ---- panels ---- */
[data-testid="stFileUploader"], [data-testid="stDataFrame"] {{
  background: var(--surface); border: 1px solid var(--line); border-radius: 10px;
}}
[data-testid="stFileUploader"] {{ padding: .5rem .75rem; }}
[data-testid="stFileUploader"] section {{ border-radius: 8px; }}
[data-testid="stFileUploaderDropzone"] {{ background: transparent; }}

/* ---- controls ---- */
[data-testid="stFormSubmitButton"] button,
[data-testid="stDownloadButton"] button,
[data-testid="stBaseButton-primary"] {{
  background: var(--brand) !important; color: #fff !important;
  border: 1px solid var(--brand) !important;
  border-radius: 8px; font-weight: 600; letter-spacing: .02em;
  padding: .5rem 1.15rem; min-height: 44px; cursor: pointer;
  transition: background 180ms ease, box-shadow 180ms ease;
}}
[data-testid="stFormSubmitButton"] button:hover,
[data-testid="stDownloadButton"] button:hover,
[data-testid="stBaseButton-primary"]:hover {{
  background: var(--brand-dark) !important; border-color: var(--brand-dark) !important;
  box-shadow: 0 2px 8px rgba(0,90,167,.25);
}}
[data-testid="stFormSubmitButton"] button:focus-visible,
[data-testid="stDownloadButton"] button:focus-visible,
[data-testid="stBaseButton-primary"]:focus-visible,
input:focus-visible {{
  outline: 3px solid rgba(0,90,167,.45); outline-offset: 2px;
}}
/* the uploader's own "Browse files" button stays secondary */
[data-testid="stFileUploader"] button {{
  border-radius: 8px; font-weight: 600; cursor: pointer; min-height: 44px;
  border-color: var(--brand); color: var(--brand);
}}
[data-testid="stFileUploader"] button:hover {{
  background: rgba(0,90,167,.06); border-color: var(--brand-dark); color: var(--brand-dark);
}}
[data-testid="stTextInput"] input {{ border-radius: 8px; }}
[data-testid="stTextInput"] input:focus {{ border-color: var(--brand); }}

/* ---- login card ---- */
.hb-login {{
  background: var(--surface); border: 1px solid var(--line); border-top: 3px solid var(--brand);
  border-radius: 14px; padding: 2rem 2rem 1.2rem;
  box-shadow: 0 10px 30px rgba(22,23,30,.07);
}}
.hb-login h2 {{
  font-family: 'Teko','Barlow',sans-serif; font-weight: 600; font-size: 1.75rem;
  color: var(--ink); margin: .9rem 0 .1rem; line-height: 1.1;
}}
.hb-login p {{ color: var(--grey); font-size: .9rem; margin: 0 0 .4rem; }}

/* ---- footer ---- */
.hb-foot {{
  margin-top: 2.2rem; padding-top: 1rem; border-top: 1px solid var(--line);
  color: var(--grey); font-size: .78rem; display: flex; gap: .5rem; flex-wrap: wrap;
}}
.hb-foot b {{ color: var(--ink); font-weight: 600; }}

@media (prefers-reduced-motion: reduce) {{
  * {{ transition: none !important; animation: none !important; }}
}}
@media (max-width: 640px) {{
  .hb-head {{ flex-wrap: wrap; }}
  .hb-head .hb-title {{ font-size: 1.6rem; }}
}}
</style>
"""


def masthead() -> None:
    # NB: keep this HTML flush left — indented lines are parsed as a code block.
    st.markdown(
        CSS
        + '<div class="hb-head">'
        + LOGO_SVG.strip()
        + '<div><p class="hb-title">Price History Pivot</p>'
        '<p class="hb-sub">Harb Electric &middot; Tendering</p></div>'
        '<div class="hb-spacer"></div>'
        '<span class="hb-badge">Internal tool</span>'
        "</div>",
        unsafe_allow_html=True,
    )


def step(number: int, label: str) -> None:
    st.markdown(
        f'<div class="hb-step"><span class="n">{number}</span>{label}</div>',
        unsafe_allow_html=True,
    )


def footer() -> None:
    st.markdown(
        '<div class="hb-foot"><b>Harb Electric</b><span>&middot;</span>'
        "<span>Files are processed in-session and never stored.</span></div>",
        unsafe_allow_html=True,
    )


# --------------------------------------------------------------------------- auth


def check_password() -> bool:
    """Show a password prompt until the correct password is entered."""
    if st.session_state.get("authenticated"):
        return True

    try:
        expected = st.secrets.get("app_password")
    except Exception:  # no secrets.toml at all
        expected = None
    st.markdown(CSS, unsafe_allow_html=True)
    if not expected:
        st.error(
            "No app password is configured. Add `app_password = \"...\"` to the app's "
            "secrets (Streamlit Cloud → Settings → Secrets, or `.streamlit/secrets.toml` "
            "when running locally)."
        )
        return False

    _, mid, _ = st.columns([1, 1.1, 1])
    with mid:
        st.markdown(
            f'<div class="hb-login">{LOGO_SVG}'
            "<h2>Price History Pivot</h2>"
            "<p>Harb Electric internal tool. Enter the access password to continue.</p>",
            unsafe_allow_html=True,
        )
        with st.form("login"):
            pwd = st.text_input(
                "Password", type="password", autocomplete="current-password",
                placeholder="Access password",
            )
            submitted = st.form_submit_button(
                "Sign in", use_container_width=True, type="primary"
            )
        st.markdown("</div>", unsafe_allow_html=True)

        if submitted:
            if hmac.compare_digest(pwd, str(expected)):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password. Please try again.")
    return False


# ---------------------------------------------------------------------- parsing


def norm(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str("" if value is None else value).lower())


def map_headers(header_row: list) -> dict:
    """Map column positions to our field names."""
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        n = norm(cell)
        if not n:
            continue
        for field, aliases in FIELDS.items():
            if field not in mapping and n in aliases:
                mapping[field] = i
    # Fall back to a plain "Net Price" column if "OC Net Price" is absent.
    if "price" not in mapping:
        for i, cell in enumerate(header_row):
            if norm(cell) == "netprice":
                mapping["price"] = i
                break
    return mapping


def find_header(rows: list[list]) -> tuple[int, dict] | None:
    """The first row (within the first 20) giving both a No. and a Posting Date column."""
    for r in range(min(len(rows), 20)):
        mapping = map_headers(rows[r])
        if "no" in mapping and "date" in mapping:
            return r, mapping
    return None


def to_date(value):
    """Excel serials, datetimes and day-first strings -> date. Otherwise None."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()
    if isinstance(value, (int, float)):
        serial = float(value)
        if serial < 1 or serial > 2958465:
            return None
        # 1900 date system, including Excel's phantom 1900-02-29.
        days = int(serial) - (0 if serial < 61 else 1)
        return date(1899, 12, 31) + timedelta(days=days)

    text = str(value).strip()
    if not text:
        return None
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", text)
    if m:
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000 if y < 70 else 1900
        try:
            return date(y, mth, d)
        except ValueError:
            return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def to_num(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[\s,']", "", str(value).strip())
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def build_pivot(rows: list[list], include_desc: bool) -> dict:
    found = find_header(rows)
    if not found:
        raise ValueError('Could not find a header row containing both "No." and "Posting Date".')
    header_row, mapping = found

    missing = [f for f in ("no", "qty", "price", "curr", "date") if f not in mapping]
    if missing:
        raise ValueError("Missing column(s): " + ", ".join(LABELS[f] for f in missing))

    groups: dict[str, dict] = {}
    order: list[str] = []
    skipped = 0

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        raw_no = cell(row, mapping["no"])
        no = "" if raw_no is None or pd.isna(raw_no) else str(raw_no).strip()
        if no.endswith(".0") and str(raw_no).replace(".", "", 1).isdigit():
            no = no[:-2]  # numeric item codes read back as floats
        if not no:
            if any(c is not None and not pd.isna(c) and str(c).strip() for c in row):
                skipped += 1
            continue

        key = no.upper()
        if key not in groups:
            groups[key] = {"no": no, "desc": "", "entries": []}
            order.append(key)
        group = groups[key]
        if include_desc and not group["desc"] and "desc" in mapping:
            d = cell(row, mapping["desc"])
            group["desc"] = "" if d is None or pd.isna(d) else str(d).strip()

        curr = cell(row, mapping["curr"])
        group["entries"].append(
            {
                "date": to_date(cell(row, mapping["date"])),
                "qty": to_num(cell(row, mapping["qty"])),
                "price": to_num(cell(row, mapping["price"])),
                "curr": "" if curr is None or pd.isna(curr) else str(curr).strip(),
                "seq": r,
            }
        )

    if not order:
        raise ValueError("No data rows found under the header.")

    # Newest first; undated lines sink to the bottom, ties keep original file order.
    max_dup = 0
    for key in order:
        entries = groups[key]["entries"]
        entries.sort(key=lambda e: (e["date"] is not None, e["date"] or date.min, -e["seq"]), reverse=True)
        max_dup = max(max_dup, len(entries))

    header = ["No."] + (["Description"] if include_desc else [])
    for _ in range(max_dup):
        header += BLOCK

    body = []
    for key in order:
        group = groups[key]
        out = [group["no"]] + ([group["desc"]] if include_desc else [])
        for i in range(max_dup):
            if i < len(group["entries"]):
                e = group["entries"][i]
                out += [i + 1, e["date"], e["qty"], e["price"], e["curr"]]
            else:
                out += [None] * 5
        body.append(out)

    return {
        "header": header,
        "body": body,
        "max_dup": max_dup,
        "items": len(order),
        "lines": sum(len(groups[k]["entries"]) for k in order),
        "skipped": skipped,
        "include_desc": include_desc,
    }


# ----------------------------------------------------------------------- export


def to_excel(pivot: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Price History"

    ws.append(pivot["header"])
    for c in range(1, len(pivot["header"]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for row in pivot["body"]:
        ws.append(row)

    first_block = 2 if pivot["include_desc"] else 1  # 0-based count of leading columns
    widths = [18] + ([40] if pivot["include_desc"] else [])
    widths += [4, 12, 9, 12, 7] * pivot["max_dup"]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    blue_fill = PatternFill("solid", start_color=R_BLUE, end_color=R_BLUE)
    blue_font = Font(bold=True, color="FFFFFF")

    for b in range(pivot["max_dup"]):
        r_col = first_block + b * 5 + 1     # 1-based R column of this block
        date_col = r_col + 1
        for r in range(2, len(pivot["body"]) + 2):
            ws.cell(row=r, column=date_col).number_format = "DD/MM/YYYY"
            r_cell = ws.cell(row=r, column=r_col)
            if r_cell.value is not None:    # only filled R cells, not the padding
                r_cell.fill = blue_fill
                r_cell.font = blue_font
                r_cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=2, column=first_block + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_frame(pivot: dict) -> pd.DataFrame:
    """Preview frame with de-duplicated column labels and dates as dd/mm/yyyy text."""
    labels = []
    counts: dict[str, int] = {}
    for name in pivot["header"]:
        counts[name] = counts.get(name, 0) + 1
        labels.append(name if counts[name] == 1 else f"{name}.{counts[name]}")

    rows = [
        [v.strftime("%d/%m/%Y") if isinstance(v, date) else ("" if v is None else v) for v in row]
        for row in pivot["body"]
    ]
    return pd.DataFrame(rows, columns=labels)


def style_frame(df: pd.DataFrame, pivot: dict):
    """Blue-fill the R cells that hold an occurrence number, matching the export."""
    first_block = 2 if pivot["include_desc"] else 1
    r_cols = [df.columns[first_block + b * 5] for b in range(pivot["max_dup"])]
    css = f"background-color: #{R_BLUE}; color: white; font-weight: 600;"

    def paint(col: pd.Series):
        return [css if v != "" else "" for v in col]

    return df.style.apply(paint, subset=r_cols)


# -------------------------------------------------------------------------- app


def stat_cards(pivot: dict) -> None:
    cards = [
        (pivot["items"], "Unique items"),
        (pivot["lines"], "Source lines"),
        (pivot["max_dup"], "Max transactions / item"),
    ]
    st.markdown(
        '<div class="hb-stats">'
        + "".join(f'<div class="hb-stat"><div class="v">{v:,}</div>'
                  f'<div class="k">{k}</div></div>' for v, k in cards)
        + "</div>",
        unsafe_allow_html=True,
    )


def main() -> None:
    masthead()

    step(1, "Upload the sales-lines export")
    uploaded = st.file_uploader(
        "Excel or CSV file",
        type=["xlsx", "xlsm", "xls", "csv"],
        label_visibility="collapsed",
        help="Expected columns: No., Description, Document No., Customer Name, Quantity, "
        "Pre-Discount Price, Discount %, Net Price, OC Net Price, Currency, Posting Date",
    )
    if not uploaded:
        st.caption(
            "Accepted: .xlsx, .xlsm, .xls, .csv — the header row does not have to be the first "
            "row. Files are processed in this session only and are never stored."
        )
        footer()
        return

    data = uploaded.getvalue()
    try:
        if uploaded.name.lower().endswith(".csv"):
            sheet_rows = {"CSV": pd.read_csv(io.BytesIO(data), header=None, dtype=object)}
        else:
            book = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None, dtype=object)
            sheet_rows = book
    except Exception as exc:  # noqa: BLE001 — surface any reader failure to the user
        st.error(f"Could not read that file: {exc}")
        footer()
        return

    step(2, "Choose the sheet")
    col1, col2 = st.columns([2, 1])
    with col1:
        sheet = st.selectbox(
            "Sheet", list(sheet_rows.keys()),
            help="Pick the tab holding the sales lines.",
        )
    with col2:
        include_desc = st.checkbox(
            "Include a Description column",
            help="Adds the item description beside the item number.",
        )

    rows = sheet_rows[sheet].where(pd.notna(sheet_rows[sheet]), None).values.tolist()

    try:
        pivot = build_pivot(rows, include_desc)
    except ValueError as exc:
        st.error(str(exc))
        footer()
        return

    step(3, "Review and export")
    stat_cards(pivot)
    if pivot["skipped"]:
        st.warning(
            f"{pivot['skipped']} row(s) were skipped because they had no No. value. "
            "Check for subtotal or note rows in the source file."
        )

    st.dataframe(style_frame(to_frame(pivot), pivot), use_container_width=True, hide_index=True)
    st.caption(
        "Each block reads newest → oldest. Blue **R** cells number the transactions per item; "
        "blank blocks mean that item has fewer transactions."
    )

    st.download_button(
        "Download Excel",
        data=to_excel(pivot),
        file_name=f"price-history-pivot-{date.today():%Y-%m-%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    footer()


if __name__ == "__main__" and check_password():
    main()
