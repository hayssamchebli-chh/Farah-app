"""Pivot engine shared by every page of the app.

Reads a flat transaction export and returns one row per item, with each
transaction laid out newest -> oldest as repeated [R, Date, Qty, Price, Curr.]
blocks. Column names differ per source document, so each page passes a
`Profile` describing the header aliases to look for.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BLOCK = ["R", "Date", "Qty", "Price", "Curr."]
REQUIRED = ("no", "qty", "price", "curr", "date")


@dataclass(frozen=True)
class Profile:
    """Everything that differs between one source document and another."""

    key: str
    title: str
    blurb: str
    # field -> accepted header names, matched after normalising
    fields: dict[str, list[str]]
    # field -> the name shown in error messages
    labels: dict[str, str]
    expected: str                                  # help text listing the columns
    file_stem: str
    # optional: let the user pick which column supplies Price
    price_choices: dict[str, list[str]] = field(default_factory=dict)
    price_fallback: list[str] = field(default_factory=list)

    def with_price(self, aliases: list[str]) -> Profile:
        """Copy of this profile with Price read from `aliases` instead."""
        return Profile(**{**self.__dict__, "fields": {**self.fields, "price": aliases}})


# --------------------------------------------------------------------------- cells


def norm(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str("" if value is None else value).lower())


def to_date(value):
    """Excel serials, datetimes and day-first strings -> date. Otherwise None."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()
    if isinstance(value, (int, float)):
        serial = float(value)
        if serial < 1 or serial > 2958465:
            return None
        # 1900 date system, including Excel's phantom 1900-02-29.
        days = int(serial) - (0 if serial < 61 else 1)
        return date(1899, 12, 31) + timedelta(days=days)

    text = str(value).strip()
    if not text:
        return None
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", text)
    if m:
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000 if y < 70 else 1900
        try:
            return date(y, mth, d)
        except ValueError:
            return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def to_num(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[\s,']", "", str(value).strip())
    if text.startswith("(") and text.endswith(")"):       # (1,234) = negative
        text = "-" + text[1:-1]
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# -------------------------------------------------------------------------- header


def map_headers(header_row: list, profile: Profile) -> dict:
    """Map column positions to field names."""
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        n = norm(cell)
        if not n:
            continue
        for name, aliases in profile.fields.items():
            if name not in mapping and n in aliases:
                mapping[name] = i
    if "price" not in mapping and profile.price_fallback:
        for i, cell in enumerate(header_row):
            if norm(cell) in profile.price_fallback:
                mapping["price"] = i
                break
    return mapping


def find_header(
    rows: list[list], profile: Profile, must_have: tuple[str, ...] = ("no", "date")
) -> tuple[int, dict] | None:
    """First row (within the first 20) that carries all of `must_have`."""
    for r in range(min(len(rows), 20)):
        mapping = map_headers(rows[r], profile)
        if all(f in mapping for f in must_have):
            return r, mapping
    return None


# --------------------------------------------------------------------------- pivot


def build_pivot(rows: list[list], profile: Profile, include_desc: bool) -> dict:
    found = find_header(rows, profile, ("no", "date"))
    if not found:
        raise ValueError(
            f'Could not find a header row containing both "{profile.labels["no"]}" '
            f'and "{profile.labels["date"]}".'
        )
    header_row, mapping = found

    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError("Missing column(s): " + ", ".join(profile.labels[f] for f in missing))

    groups: dict[str, dict] = {}
    order: list[str] = []
    skipped = 0

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        raw_no = cell(row, mapping["no"])
        no = "" if raw_no is None or pd.isna(raw_no) else str(raw_no).strip()
        if no.endswith(".0") and str(raw_no).replace(".", "", 1).isdigit():
            no = no[:-2]  # numeric item codes read back as floats
        if not no:
            if any(c is not None and not pd.isna(c) and str(c).strip() for c in row):
                skipped += 1
            continue

        key = no.upper()
        if key not in groups:
            groups[key] = {"no": no, "desc": "", "entries": []}
            order.append(key)
        group = groups[key]
        if include_desc and not group["desc"] and "desc" in mapping:
            d = cell(row, mapping["desc"])
            group["desc"] = "" if d is None or pd.isna(d) else str(d).strip()

        curr = cell(row, mapping["curr"])
        group["entries"].append(
            {
                "date": to_date(cell(row, mapping["date"])),
                "qty": to_num(cell(row, mapping["qty"])),
                "price": to_num(cell(row, mapping["price"])),
                "curr": "" if curr is None or pd.isna(curr) else str(curr).strip(),
                "seq": r,
            }
        )

    if not order:
        raise ValueError("No data rows found under the header.")

    # Newest first; undated lines sink to the bottom, ties keep original file order.
    max_dup = 0
    for key in order:
        entries = groups[key]["entries"]
        entries.sort(
            key=lambda e: (e["date"] is not None, e["date"] or date.min, -e["seq"]),
            reverse=True,
        )
        max_dup = max(max_dup, len(entries))

    header = [profile.labels["no"]] + (["Description"] if include_desc else [])
    for _ in range(max_dup):
        header += BLOCK

    body = []
    for key in order:
        group = groups[key]
        out = [group["no"]] + ([group["desc"]] if include_desc else [])
        for i in range(max_dup):
            if i < len(group["entries"]):
                e = group["entries"][i]
                out += [i + 1, e["date"], e["qty"], e["price"], e["curr"]]
            else:
                out += [None] * 5
        body.append(out)

    return {
        "header": header,
        "body": body,
        "max_dup": max_dup,
        "items": len(order),
        "lines": sum(len(groups[k]["entries"]) for k in order),
        "skipped": skipped,
        "include_desc": include_desc,
    }


# ------------------------------------------------------------------ receipt check

RECEIPT_REQUIRED = ("no", "qty", "recv")


def build_receipt(rows: list[list], profile: Profile, extras: bool = False) -> dict:
    """Collapse repeated Item No. lines into one row per item.

    Quantity (what the purchase order expects) and Qty. to Receive (what the
    warehouse is actually receiving) are summed per item, and Difference is
    warehouse minus PO: negative = short, positive = over. The line count and
    Over-Receipt Quantity are counted for the summary but stay out of the table.
    """
    found = find_header(rows, profile, ("no", "qty"))
    if not found:
        raise ValueError(
            f'Could not find a header row containing both "{profile.labels["no"]}" '
            f'and "{profile.labels["qty"]}".'
        )
    header_row, mapping = found

    missing = [f for f in RECEIPT_REQUIRED if f not in mapping]
    if missing:
        raise ValueError("Missing column(s): " + ", ".join(profile.labels[f] for f in missing))

    groups: dict[str, dict] = {}
    order: list[str] = []
    skipped = 0

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def text(row, field) -> str:
        if field not in mapping:
            return ""
        v = cell(row, mapping[field])
        return "" if v is None or pd.isna(v) else str(v).strip()

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        raw_no = cell(row, mapping["no"])
        no = "" if raw_no is None or pd.isna(raw_no) else str(raw_no).strip()
        if no.endswith(".0") and str(raw_no).replace(".", "", 1).isdigit():
            no = no[:-2]
        if not no:
            if any(c is not None and not pd.isna(c) and str(c).strip() for c in row):
                skipped += 1
            continue

        key = no.upper()
        if key not in groups:
            groups[key] = {
                "no": no, "desc": "", "qty": 0.0, "recv": 0.0,
                "lines": 0, "bins": [], "sources": [],
            }
            order.append(key)
        g = groups[key]
        g["lines"] += 1
        if not g["desc"]:
            g["desc"] = text(row, "desc")
        for field in ("qty", "recv"):
            value = to_num(cell(row, mapping[field])) if field in mapping else None
            g[field] += value or 0.0
        for field, bucket in (("bin", "bins"), ("source", "sources")):
            value = text(row, field)
            if value and value not in g[bucket]:
                g[bucket].append(value)

    if not order:
        raise ValueError("No data rows found under the header.")

    header = [profile.labels["no"], "Description"]
    if extras:
        header += ["Source No.", "Bin Code"]
    header += [profile.labels["qty"], profile.labels["recv"], "Difference"]
    diff_col = header.index("Difference")

    body, short, over_count, balanced = [], 0, 0, 0
    for key in order:
        g = groups[key]
        diff = g["recv"] - g["qty"]
        if diff < 0:
            short += 1
        elif diff > 0:
            over_count += 1
        else:
            balanced += 1
        out = [g["no"], g["desc"]]
        if extras:
            out += [", ".join(g["sources"]), ", ".join(g["bins"])]
        out += [g["qty"], g["recv"], diff]
        body.append(out)

    return {
        "header": header,
        "body": body,
        "diff_col": diff_col,
        "items": len(order),
        "lines": sum(groups[k]["lines"] for k in order),
        "combined": sum(1 for k in order if groups[k]["lines"] > 1),
        "short": short,
        "over": over_count,
        "balanced": balanced,
        "skipped": skipped,
    }


def receipt_to_excel(receipt: dict, neg: tuple[str, str], pos: tuple[str, str]) -> bytes:
    """Workbook with the Difference column shaded red (short) or green (over)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Receipt"

    ws.append(receipt["header"])
    for c in range(1, len(receipt["header"]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for row in receipt["body"]:
        ws.append(row)

    widths = []
    for name in receipt["header"]:
        widths.append(46 if name == "Description" else (22 if name in ("Source No.", "Bin Code") else 15))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    neg_fill = PatternFill("solid", start_color=neg[0].lstrip("#").upper(),
                           end_color=neg[0].lstrip("#").upper())
    pos_fill = PatternFill("solid", start_color=pos[0].lstrip("#").upper(),
                           end_color=pos[0].lstrip("#").upper())
    neg_font = Font(bold=True, color=neg[1].lstrip("#").upper())
    pos_font = Font(bold=True, color=pos[1].lstrip("#").upper())

    diff_col = receipt["diff_col"] + 1  # 1-based
    for r in range(2, len(receipt["body"]) + 2):
        cell = ws.cell(row=r, column=diff_col)
        # the sign is spelled out in the number itself, so colour is never the
        # only thing carrying the meaning
        cell.number_format = "+#,##0.###;-#,##0.###;0"
        cell.alignment = Alignment(horizontal="center")
        value = cell.value or 0
        if value < 0:
            cell.fill, cell.font = neg_fill, neg_font
        elif value > 0:
            cell.fill, cell.font = pos_fill, pos_font

    ws.freeze_panes = ws.cell(row=2, column=3)
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def receipt_frame(receipt: dict) -> pd.DataFrame:
    return pd.DataFrame(receipt["body"], columns=receipt["header"])


def style_receipt(df: pd.DataFrame, receipt: dict, neg: tuple[str, str], pos: tuple[str, str]):
    """Shade the Difference column: red when short, green when over."""
    diff = df.columns[receipt["diff_col"]]

    def paint(col: pd.Series):
        out = []
        for v in col:
            if v < 0:
                out.append(f"background-color: {neg[0]}; color: {neg[1]}; font-weight: 700;")
            elif v > 0:
                out.append(f"background-color: {pos[0]}; color: {pos[1]}; font-weight: 700;")
            else:
                out.append("")
        return out

    styler = df.style.apply(paint, subset=[diff])
    return styler.format({diff: lambda v: f"{v:+,g}" if v else "0"})


# -------------------------------------------------------------------------- output


def to_excel(pivot: dict, accent: str, sheet_title: str = "Price History") -> bytes:
    accent = accent.lstrip("#").upper()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    ws.append(pivot["header"])
    for c in range(1, len(pivot["header"]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for row in pivot["body"]:
        ws.append(row)

    first_block = 2 if pivot["include_desc"] else 1  # count of leading columns
    widths = [18] + ([40] if pivot["include_desc"] else [])
    widths += [4, 12, 9, 12, 7] * pivot["max_dup"]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fill = PatternFill("solid", start_color=accent, end_color=accent)
    white_bold = Font(bold=True, color="FFFFFF")

    for b in range(pivot["max_dup"]):
        r_col = first_block + b * 5 + 1     # 1-based R column of this block
        date_col = r_col + 1
        for r in range(2, len(pivot["body"]) + 2):
            ws.cell(row=r, column=date_col).number_format = "DD/MM/YYYY"
            r_cell = ws.cell(row=r, column=r_col)
            if r_cell.value is not None:    # only filled R cells, not the padding
                r_cell.fill = fill
                r_cell.font = white_bold
                r_cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=2, column=first_block + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_frame(pivot: dict) -> pd.DataFrame:
    """Preview frame with de-duplicated column labels and dates as dd/mm/yyyy text."""
    labels = []
    counts: dict[str, int] = {}
    for name in pivot["header"]:
        counts[name] = counts.get(name, 0) + 1
        labels.append(name if counts[name] == 1 else f"{name}.{counts[name]}")

    rows = [
        [v.strftime("%d/%m/%Y") if isinstance(v, date) else ("" if v is None else v) for v in row]
        for row in pivot["body"]
    ]
    return pd.DataFrame(rows, columns=labels)


def style_frame(df: pd.DataFrame, pivot: dict, accent: str):
    """Shade the R cells that hold an occurrence number, matching the export."""
    first_block = 2 if pivot["include_desc"] else 1
    r_cols = [df.columns[first_block + b * 5] for b in range(pivot["max_dup"])]
    css = f"background-color: {accent}; color: white; font-weight: 600;"

    def paint(col: pd.Series):
        return [css if v != "" else "" for v in col]

    return df.style.apply(paint, subset=r_cols)


# ------------------------------------------------------------------------ profiles

SALES = Profile(
    key="sales",
    title="Selling Prices",
    blurb="Sales lines — the last price each item sold at, newest first.",
    fields={
        "no": ["no", "nos", "itemno", "item", "itemnumber", "partno", "partnumber"],
        "desc": ["description", "itemdescription", "desc"],
        "qty": ["quantity", "qty"],
        "price": ["ocnetprice", "netpriceoc", "ocnet", "priceoc"],
        "curr": ["currency", "curr", "currencycode"],
        "date": ["postingdate", "postingdt", "date", "documentdate"],
    },
    labels={
        "no": "No.",
        "qty": "Quantity",
        "price": "OC Net Price",
        "curr": "Currency",
        "date": "Posting Date",
    },
    expected=(
        "No., Description, Document No., Customer Name, Quantity, Pre-Discount Price, "
        "Discount %, Net Price, OC Net Price, Currency, Posting Date"
    ),
    file_stem="selling-price-history",
    price_fallback=["netprice"],
)

PURCHASES = Profile(
    key="purchases",
    title="Purchase Prices",
    blurb="Vendor purchase lines — what each item last cost, newest first.",
    fields={
        "no": ["itemno", "itemnumber", "item", "no", "partno", "partnumber"],
        "desc": ["description", "itemdescription", "desc"],
        "qty": ["qty", "quantity", "qtys"],
        "price": ["unitcost", "unitprice", "cost"],
        "curr": ["currency", "curr", "currencycode"],
        "date": ["date", "documentdate", "postingdate", "postingdt"],
    },
    labels={
        "no": "Item No",
        "qty": "QTY.",
        "price": "Unit Cost",
        "curr": "Currency",
        "date": "Date",
    },
    expected=(
        "Document No, Vendor, Date, Item No, Description, Unit Cost, QTY., "
        "Discount Amount, UOM, Discount, Amount, Amount Including VAT, Currency, Shipment No"
    ),
    file_stem="purchase-price-history",
    # Unit Cost is the default: it is per-unit, so prices stay comparable across
    # lines with different quantities. Amount is a line total.
    price_choices={
        "Unit Cost": ["unitcost", "unitprice", "cost"],
        "Amount": ["amount"],
        "Amount Including VAT": ["amountincludingvat", "amountinclvat"],
    },
    price_fallback=["amount"],
)

RECEIPT = Profile(
    key="receipt",
    title="Warehouse Receipt",
    blurb=(
        "Warehouse receipt lines — one row per item, with what the warehouse is "
        "receiving checked against what the order expects."
    ),
    fields={
        "no": ["itemno", "itemnumber", "item", "no"],
        "desc": ["description", "itemdescription", "desc"],
        "qty": ["quantity", "qty"],
        "recv": ["qtytoreceive", "quantitytoreceive", "qtytorecieve", "qtytoreceived"],
        "over": ["overreceiptquantity", "overreceiptqty", "overreceipt"],
        "bin": ["bincode", "bin"],
        "source": ["sourceno", "sourcenumber"],
        "srcdoc": ["sourcedocument"],
    },
    labels={
        "no": "Item No.",
        "qty": "Quantity",
        "recv": "Qty. to Receive",
        "over": "Over-Receipt Quantity",
        "bin": "Bin Code",
        "source": "Source No.",
    },
    expected=(
        "Source Document, Source No., Item No., Description, Bin Code, Quantity, "
        "Qty. to Receive, Over-Receipt Quantity"
    ),
    file_stem="warehouse-receipt-check",
)
